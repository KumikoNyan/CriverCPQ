from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from .models import *
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse
from django.contrib.auth.hashers import make_password, check_password
from collections import defaultdict
import json
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
from django.core.serializers.json import DjangoJSONEncoder

# access control via session-based checker - a simple logic to check if superuser or not
def is_logged_in(request):
    return request.session.get("account_id", None) is not None

def is_superuser(request):
    return request.session.get("is_superuser", False)

def login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        try:
            account = Account.objects.get(account_name=username)

            if check_password(password, account.account_password):
                request.session["account_id"] = account.account_id
                request.session["account_level"] = account.account_level
                request.session["is_superuser"] = account.is_superuser

                return redirect('quotation_list' if account.account_level == 'regular' else 'product_list')
            else:
                return render(request, 'cpq_app/login.html', {'error': 'Incorrect password'})
        except Account.DoesNotExist:
            return render(request, 'cpq_app/login.html', {'error': 'Account does not exist'})

    return render(request, 'cpq_app/login.html')

def logout(request):
    if not is_logged_in(request):
        return redirect('login')
        
    request.session.flush()
    return redirect('login')

def access_error(request):
    return render(request, 'cpq_app/access_error.html')

def create_account(request):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')
    
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        account_level = request.POST.get("account_level")
        is_superuser = (account_level == "superuser")

        if Account.objects.filter(account_name=username).exists():
            return render(request, 'cpq_app/create_account.html', {'error': 'Username already exists.'})

        # hashed password for security reasons
        hashed_password = make_password(password)

        Account.objects.create(
            account_name=username,
            account_password=hashed_password,
            account_level=account_level,
            is_superuser=is_superuser
        )

        return redirect('login')  # or wherever you want to redirect

    return render(request, 'cpq_app/create_account.html')

# TEMPORARY superuser seeding view DO NOT ENABLE only use ONCE
'''def init_create_superuser(request):
    if Account.objects.filter(account_level='superuser').exists():
        return JsonResponse({"message": "Superuser already exists."})

    superuser = Account.objects.create(
        account_name="admin",
        account_password=make_password("admin123"),
        account_level="superuser"
    )
    return JsonResponse({"message": "Superuser created.", "username": superuser.account_name})'''

# change welcome message later or remove if not needed
def index(request):
    return JsonResponse({"message": "Welcome Bitch!"})

def view_version(request, quotation_id):
    if not is_logged_in(request):
        return redirect('login')

    customers = Customer.objects.all()
    suppliers = Supplier.objects.all()
    materials_objects = Material.objects.all()
    
    materials = [
        {
            'material_id': material.material_id,
            'material_name': material.material_name,
            'material_supplier': material.supplier.supplier_name,
            'material_type': material.material_type
        }
        for material in materials_objects
    ]

    quotation = Quotation.objects.get(quotation_id=quotation_id)
    quotation_data = {
        'quotation_id': quotation.quotation_id,
        'customer_name': quotation.customer.customer_name,
        'customer_address': quotation.customer.customer_address,
        'customer_email': quotation.customer.customer_email,
        'customer_mobile': quotation.customer.customer_mobile,
        'customer_id': quotation.customer.customer_id,
        'project': quotation.project,
        'date_created': quotation.date_created,
        'version_number': quotation.version_number,
        'quotation_id': quotation.quotation_id,
        'items': [
            {
                'item_id': item.item_id,
                'product': item.product.product_id,
                'supplier': item.product.supplier.supplier_id,
                'product_name': item.product.product_name,
                'supplier': item.product.supplier.supplier_id,
                'item_quantity': item.item_quantity,
                'item_margin': item.product_margin,
                'item_labor': item.product_labor,
                'item_height': item.item_height,
                'item_width': item.item_width,
                'glass_finish': item.glass_finish,
                'aluminum_finish': item.aluminum_finish,
                'excluded_materials': '-'.join(item.excluded_materials.split('-')) if item.excluded_materials else '',
                'item_label': item.item_label,
                'item_materials': '|'.join([
                    f"{im.material.material_id}/{im.finish}/{im.quantity}"
                    for im in ItemMaterial.objects.filter(item=item)
                ]),
            } for item in QuotationItem.objects.filter(quotation=quotation)
        ],
    }
    print(quotation_data)

    if quotation_data['items']:
        quotation_data['quotation_margin'] = quotation_data['items'][0]['item_margin']
        quotation_data['quotation_labor'] = quotation_data['items'][0]['item_labor']
    else:
        quotation_data['quotation_margin'] = 0
        quotation_data['quotation_labor'] = 0

    supplier_products = defaultdict(list)
    products = Product.objects.select_related('supplier').all()
    for product in products:
        supplier_name = product.supplier.supplier_name
        supplier_products[supplier_name].append({
            'id': product.product_id,
            'name': product.product_name
        })

    return render(
        request,
        'cpq_app/view_version.html',
        {
            'customers': customers,
            'suppliers': suppliers,
            'quotation': quotation_data,
            'item_count': len(quotation_data['items']),
            'supplier_products': supplier_products,
            'materials': json.dumps(materials, cls=DjangoJSONEncoder)
        }
    )

# quotation views
@csrf_exempt
def quotation_list(request):
    if not is_logged_in(request):
        return redirect('login')

    active_quotations = Quotation.objects.filter(is_active_version=True)
    quotation_data = []

    for quotation in active_quotations:
        # Fetch all previous versions (non-active ones) in the same group
        previous_versions = Quotation.objects.filter(
            quotation_group_id=quotation.quotation_group_id,
            is_active_version=False
        ).values('version_number', 'quotation_id')

        temp = {
            'customer_name': quotation.customer.customer_name,
            'project': quotation.project,
            'date_created': quotation.date_created,
            'version_number': quotation.version_number,
            'quotation_group_id': quotation.quotation_group_id,
            'quotation_id': quotation.quotation_id,
            'previous_versions': list(previous_versions)  # convert queryset to list of dicts
        }

        quotation_data.append(temp)


    if request.method == "POST":
        response = {}
        response['status'] = True

        if request.POST.get("action") == "delete":
            quotation_id = request.POST.get("quotation_id")
            quotation_object = Quotation.objects.get(quotation_id=quotation_id)
            quotation_object.delete()
            response['url'] = reverse('quotation_list')
            print(response)
            return JsonResponse(response)
    return render(request, 'cpq_app/quotation_list.html', {"quotations": quotation_data})

def quotation_detail(request, quotation_id):
    if not is_logged_in(request):
        return redirect('login')

    customers = Customer.objects.all()
    suppliers = Supplier.objects.all()
    materials_objects = Material.objects.all()
    
    materials = [
        {
            'material_id': material.material_id,
            'material_name': material.material_name,
            'material_supplier': material.supplier.supplier_name,
            'material_type': material.material_type
        }
        for material in materials_objects
    ]

    if request.method == "POST":
        response = {'status': True}

       
        old_quotation = get_object_or_404(Quotation, pk=quotation_id)
        customer = get_object_or_404(Customer, customer_id=request.POST.get("customer_id"))
        project = request.POST.get("project")

        Quotation.objects.filter(
            customer=old_quotation.customer,
            project=old_quotation.project
        ).update(is_active_version=False)

        latest_version = Quotation.objects.filter(
            customer=old_quotation.customer,
            project=old_quotation.project
        ).aggregate(models.Max('version_number'))['version_number__max'] or 0

        new_quotation = Quotation.objects.create(
            customer=customer,
            project=project,
            version_number=latest_version + 1,
            is_active_version=True,
            quotation_group_id=old_quotation.quotation_group_id
        )

        for item in json.loads(request.POST.get('item_data')):
            product = Product.objects.get(product_id=item['product_id'])

            new_item = QuotationItem.objects.create(
                quotation=new_quotation,
                product=product,
                item_quantity=int(item['item_quantity']),
                product_margin=int(item['submit_margin']),
                product_labor=int(item['submit_labor']),
                item_height=float(item['item_height']),
                item_width=float(item['item_width']),
                glass_finish=item['glass_finish'],
                aluminum_finish=item['aluminum_finish'],
                excluded_materials=item.get('excluded_materials', ''),
                item_label=item['item_label']
            )

            additional_materials = json.loads(item['additional_materials'])
            if additional_materials:
                for material in additional_materials:
                    material_obj = Material.objects.get(material_id=material['material_id'])
                    item_material = ItemMaterial.objects.create(item=new_item, material=material_obj, finish=material['finish'], quantity=material['quantity'])

        response['url'] = reverse('quotation_list')
        return JsonResponse(response)

    quotation = Quotation.objects.get(quotation_id=quotation_id)
    quotation_data = {
        'quotation_id': quotation.quotation_id,
        'customer_name': quotation.customer.customer_name,
        'customer_address': quotation.customer.customer_address,
        'customer_email': quotation.customer.customer_email,
        'customer_mobile': quotation.customer.customer_mobile,
        'customer_id': quotation.customer.customer_id,
        'project': quotation.project,
        'date_created': quotation.date_created,
        'version_number': quotation.version_number,
        'quotation_id': quotation.quotation_id,
        'items': [
            {
                'item_id': item.item_id,
                'product': item.product.product_id,
                'supplier': item.product.supplier.supplier_id,
                'product_name': item.product.product_name,
                'supplier': item.product.supplier.supplier_id,
                'item_quantity': item.item_quantity,
                'item_margin': item.product_margin,
                'item_labor': item.product_labor,
                'item_height': item.item_height,
                'item_width': item.item_width,
                'glass_finish': item.glass_finish,
                'aluminum_finish': item.aluminum_finish,
                'excluded_materials': '-'.join(item.excluded_materials.split('-')) if item.excluded_materials else '',
                'item_label': item.item_label,
                'item_materials': '|'.join([
                    f"{im.material.material_id}/{im.finish}/{im.quantity}"
                    for im in ItemMaterial.objects.filter(item=item)
                ]),
            } for item in QuotationItem.objects.filter(quotation=quotation)
        ],
    }
    print(quotation_data)

    if quotation_data['items']:
        quotation_data['quotation_margin'] = quotation_data['items'][0]['item_margin']
        quotation_data['quotation_labor'] = quotation_data['items'][0]['item_labor']
    else:
        quotation_data['quotation_margin'] = 0
        quotation_data['quotation_labor'] = 0

    supplier_products = defaultdict(list)
    products = Product.objects.select_related('supplier').all()
    for product in products:
        supplier_name = product.supplier.supplier_name
        supplier_products[supplier_name].append({
            'id': product.product_id,
            'name': product.product_name
        })

    return render(
        request,
        'cpq_app/quotation_detail.html',
        {
            'customers': customers,
            'suppliers': suppliers,
            'quotation': quotation_data,
            'item_count': len(quotation_data['items']),
            'supplier_products': supplier_products,
            'materials': json.dumps(materials, cls=DjangoJSONEncoder)
        }
    )

@csrf_exempt
def create_quotation(request):
    if not is_logged_in(request):
        return redirect('login')

    customers = Customer.objects.all()
    suppliers = Supplier.objects.all()
    materials_objects = Material.objects.all()
    
    materials = [
        {
            'material_id': material.material_id,
            'material_name': material.material_name,
            'material_supplier': material.supplier.supplier_name,
            'material_type': material.material_type
        }
        for material in materials_objects
    ]
    if request.method == "POST":
        print(request.POST)
        response = {}
        response['status'] = True
        customer = get_object_or_404(Customer, customer_id=request.POST.get("customer_id"))
        project = request.POST.get("project")
        quotation = Quotation.objects.create(
            customer=customer,
            project=project,
            version_number=1,
            is_active_version=True
        )

        for item in json.loads(request.POST.get('item_data')):
            product = Product.objects.get(product_id=item['product_id'])
            excluded_materials = item.get('excluded_materials', [])
            if excluded_materials:
                excluded_materials_str = '-'.join(excluded_materials)
            else:
                excluded_materials_str = ''
            new_item = QuotationItem.objects.create(
                quotation=quotation,
                product=product,
                item_quantity=int(item['item_quantity']),
                product_margin=int(item['submit_margin']),
                product_labor=int(item['submit_labor']),
                item_height=float(item['item_height']),
                item_width=float(item['item_width']),
                glass_finish=item['glass_finish'],
                aluminum_finish=item['aluminum_finish'],
                excluded_materials=excluded_materials_str,
                item_label=item['item_label'],
            )

            additional_materials = json.loads(item['additional_materials'])
            if additional_materials:
                for material in additional_materials:
                    material_obj = Material.objects.get(material_id=material['material_id'])
                    item_material = ItemMaterial.objects.create(item=new_item, material=material_obj, finish=material['finish'], quantity=material['quantity'])
                    print(item_material)
            print(item.get('excluded_materials', ''))
        response['url'] = reverse('quotation_list')
        return JsonResponse(response)
    return render(request, 'cpq_app/create_quotation.html', {'customers': customers, 'suppliers': suppliers, 'materials': json.dumps(materials, cls=DjangoJSONEncoder)})


# func to get quotation versions
def get_quotation_versions(request, quotation_id):
    versions = list(Quotation.objects.filter(id=quotation_id).values())
    return JsonResponse({"quotation_versions": versions})


# quotation item views
@csrf_exempt
def add_quotation_item(request, quotation_id):
    if not is_logged_in(request):
        return redirect('login')

    if request.method == "POST":
        data = json.loads(request.body)
        quotation = get_object_or_404(Quotation, pk=quotation_id)
        product = get_object_or_404(Product, pk=data['product_id'])
        
        item = QuotationItem.objects.create(
            quotation=quotation,
            product=product,
            item_quantity=data['item_quantity'],
            unit_price=data['unit_price'],
            total_price=data['item_quantity'] * data['unit_price']
        )
        
        return JsonResponse({"message": "Item added to quotation", "item_id": item.id})

def get_quotation_items(request, quotation_id):
    items = QuotationItem.objects.filter(quotation_id=quotation_id)
    items_list = [{
        "id": item.id,
        "product": item.product.product_name,
        "quantity": item.item_quantity,
        "unit_price": item.unit_price,
        "total_price": item.total_price
    } for item in items]
    return JsonResponse({"quotation_items": items_list})

# customer views
def customer_list(request):
    if not is_logged_in(request):
        return redirect('login')

    customers = Customer.objects.all()

    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("action") == "delete":
            customer_id = request.POST.get("customer_id")
            delete_customer(request)
            response['url'] = reverse('customer_list') 
            print(response)
            return JsonResponse(response)
    return render(request, 'cpq_app/customer_list.html', {"customers": customers})

def customer_detail(request, customer_id):
    if not is_logged_in(request):
        return redirect('login')

    customer_object = get_object_or_404(Customer, customer_id=customer_id)

    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("action") == "delete":
            customer_id = request.POST.get("customer_id")
            delete_customer(customer_id) # customer_id works here for some reason instead of request
            response['url'] = reverse('customer_list') 
            print(response)
            return JsonResponse(response)
        
        else:
            customer_id = request.POST.get("customer_id")
            customer_name = request.POST.get("customer_name")
            customer_address = request.POST.get("customer_address")
            customer_mobile = request.POST.get("customer_mobile")
            customer_email = request.POST.get("customer_email")

            customer = get_object_or_404(Customer, customer_id=customer_id)
            customer.customer_name = customer_name
            customer.customer_address = customer_address
            customer.customer_mobile = customer_mobile
            customer.customer_email = customer_email
            customer.save()

            response['url'] = reverse('customer_list')
            return JsonResponse(response)
    return render(request, 'cpq_app/customer_detail.html', {'customer_object': customer_object})

def add_customer(request):
    if not is_logged_in(request):
        return redirect('login')

    customers = Customer.objects.all()

    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        customer_name = request.POST.get("customer_name")
        customer_address = request.POST.get("customer_address")
        customer_mobile = request.POST.get("customer_mobile")
        customer_email = request.POST.get("customer_email")

        if Customer.objects.filter(customer_email=customer_email).exists():
            return JsonResponse({"status": False, "error": "Customer with this email already exists."}, status=400)
            
        new_customer = Customer.objects.create(customer_name=customer_name,
        customer_address=customer_address,
        customer_mobile=customer_mobile,
        customer_email=customer_email)

        response['url'] = reverse('customer_list')
        print(response)
        return JsonResponse(response)
    return render(request, 'cpq_app/add_customer.html', {'customers': customers})

def delete_customer(request):
    if request.method == "POST":
        customer_id = request.POST.get("customer_id")
        customer_object = get_object_or_404(Customer, customer_id=customer_id)
        customer_object.delete()
        return JsonResponse({"status": True, "message": "Customer deleted successfully."})
    return JsonResponse({"status": False, "message": "Invalid Request"}, status=400)

# product views
def delete_product(product_id):
    product_object = Product.objects.get(product_id=product_id)
    product_object.delete()

def product_list(request):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    products = Product.objects.all()
    suppliers = Supplier.objects.all()
    supplier_count = len(suppliers) or 0

    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("action") == "delete":
            product_id = request.POST.get("product_id")
            delete_product(product_id)
            response['url'] = reverse('product_list')
            print(response)
            return JsonResponse(response)

        elif request.POST.get("supplier_data"):
            for supplier in json.loads(request.POST.get("supplier_data")):
                print(supplier)
                supplier_operations(supplier)
            response['url'] = reverse('product_list')
            print(response)
            return JsonResponse(response)

    return render(request, 'cpq_app/product_list.html', {"products": products, 'suppliers': suppliers, 'supplier_count': supplier_count})

def product_detail(request, product_id):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    product_object = Product.objects.get(product_id=product_id)
    product_material_object = ProductMaterial.objects.filter(product=product_object)
    suppliers = Supplier.objects.all()
    supplier_count = len(suppliers) or 0
    material_data_by_suppliers = get_material_data_by_suppliers(suppliers)
    selected_supplier_data = get_material_data_by_suppliers(suppliers, supplier_id=product_object.supplier.supplier_id)
    
    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("supplier_data"):
            for supplier in json.loads(request.POST.get("supplier_data")):
                print(supplier)
                supplier_operations(supplier)
            response['url'] = reverse('product_detail', args=[product_id])
            print(response)
            return JsonResponse(response)

        else:
            product_id = request.POST.get("product_id")
            product_name = request.POST.get("product_name")
            product_category = request.POST.get("product_category")
            supplier_id = request.POST.get("supplier")
            supplier = Supplier.objects.get(supplier_id=supplier_id)

            product = Product.objects.get(product_id = product_id)
            product.product_name = product_name
            product.product_category = product_category
            product.supplier = supplier
            
            product.save()

            pm_data = json.loads(request.POST.get("pm_data"))
            
            ProductMaterial.objects.filter(product=product).delete()
            for pm in pm_data:
                material_id = pm["material_id"]
                material_quantity = pm["material_quantity"]
                material_scale = pm["material_scale"]
                scale_ratio = pm["scale_ratio"]
                scale_ratio_second = pm.get("scale_ratio_second", None)

                print(scale_ratio)

                material = get_object_or_404(Material, material_id=material_id)

                if scale_ratio_second:
                    new_pm = ProductMaterial.objects.create(product=product, material=material, material_quantity=material_quantity, scale_by_height=True, scale_by_width=True, scale_ratio=scale_ratio, scale_ratio_second=scale_ratio_second)
                elif material_scale == 'by_height':
                    new_pm = ProductMaterial.objects.create(product=product, material=material, material_quantity=material_quantity, scale_by_height=True, scale_ratio=scale_ratio)
                elif material_scale == "by_width":
                    new_pm = ProductMaterial.objects.create(product=product, material=material, material_quantity=material_quantity, scale_by_width=True, scale_ratio=scale_ratio)
                else:
                    new_pm = ProductMaterial.objects.create(product=product, material=material, material_quantity=material_quantity, scale_ratio=0)

                print(new_pm)

            response['url'] = reverse('product_list')
            print(response)
            return JsonResponse(response)
    return render(request, 'cpq_app/product_detail.html', {'product': product_object, 'product_materials': product_material_object, 'material_data': material_data_by_suppliers, 'suppliers': suppliers, 'supplier_count': supplier_count, 'selected_supplier_pm_data': selected_supplier_data})

def create_product(request):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    materials = Material.objects.all()
    suppliers = Supplier.objects.all()

    material_data_by_suppliers = get_material_data_by_suppliers(suppliers)

    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("supplier_data"):
            for supplier in json.loads(request.POST.get("supplier_data")):
                print(supplier)
                supplier_operations(supplier)
            response['url'] = reverse('create_product')
            print(response)
            return JsonResponse(response)

        else:
            product_id = request.POST.get("product_id")
            product_name = request.POST.get("product_name")
            product_category = request.POST.get("product_category")
            supplier_id = request.POST.get("supplier")
            supplier = Supplier.objects.get(supplier_id=supplier_id)

            new_product = Product.objects.create(product_name=product_name, 
            product_category=product_category, 
            supplier=supplier)

            pm_data = json.loads(request.POST.get("pm_data"))
            if pm_data:
                for pm in pm_data:
                    material_id = pm["material_id"]
                    material_quantity = pm["material_quantity"]
                    material_scale = pm["material_scale"]
                    scale_ratio = pm["scale_ratio"]
                    scale_ratio_second = pm.get("scale_ratio_second", None)

                    material = get_object_or_404(Material, material_id=material_id)

                    if scale_ratio_second:
                        new_pm = ProductMaterial.objects.create(product=new_product, material=material, material_quantity=material_quantity, scale_by_height=True, scale_by_width=True, scale_ratio=scale_ratio, scale_ratio_second=scale_ratio_second)
                    elif material_scale == 'by_height':
                        new_pm = ProductMaterial.objects.create(product=new_product, material=material, material_quantity=material_quantity, scale_by_height=True, scale_ratio=scale_ratio)
                    elif material_scale == "by_width":
                        new_pm = ProductMaterial.objects.create(product=new_product, material=material, material_quantity=material_quantity, scale_by_width=True, scale_ratio=scale_ratio)
                    else:
                        new_pm = ProductMaterial.objects.create(product=new_product, material=material, material_quantity=material_quantity, scale_ratio=0)

                    print(new_pm)

            response['url'] = reverse('product_list')
            print(response)
            return JsonResponse(response)
    return render(request, 'cpq_app/create_product.html', {"materials": materials, "suppliers": suppliers, "material_data": material_data_by_suppliers})

# material views
def material_list(request):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    materials = Material.objects.all()
    suppliers = Supplier.objects.all()
    supplier_count = len(suppliers) or 0

    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("action") == "delete":
            material_id = request.POST.get("material_id")
            delete_material(material_id)
            response['url'] = reverse('material_list')  # URL to direct is str
            print(response)
            return JsonResponse(response)
        elif request.POST.get("supplier_data"):
            for supplier in json.loads(request.POST.get("supplier_data")):
                print(supplier)
                supplier_operations(supplier)
            response['url'] = reverse('material_list')  # URL to direct is str
            print(response)
            return JsonResponse(response)
    return render(request, 'cpq_app/material_list.html', {"materials": materials, 'suppliers': suppliers, 'supplier_count': supplier_count})

def material_detail(request, material_id):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    material_object = Material.objects.get(material_id=material_id)
    finishes = MaterialFinish.objects.filter(material=material_object)
    suppliers = Supplier.objects.all()
    supplier_count = len(suppliers) or 0
    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("action") == "delete":
            material_id = request.POST.get("material_id")
            delete_material(material_id)
            response['url'] = reverse('material_list')  # URL to direct is str
            print(response)
            return JsonResponse(response)
        else:
            if request.POST.get("supplier_data"):
                for supplier in json.loads(request.POST.get("supplier_data")):
                    print(supplier)
                    supplier_operations(supplier)
                response['url'] = reverse('create_material')  # URL to direct is str
                print(response)
                return JsonResponse(response)
            else:
                material_id = request.POST.get("material_id")
                material_name = request.POST.get("material_name")
                material_cost = request.POST.get("material_cost") or 0
                material_type = request.POST.get("material_type")
                material_unit = request.POST.get("material_unit")
                supplier_id = request.POST.get("supplier")
                finish_data = json.loads(request.POST.get("finish_data", "[]"))  # Load finishes

                material = get_object_or_404(Material, material_id=material_id)
                material.material_name = material_name
                material.material_cost = material_cost
                material.material_type = material_type
                material.material_unit = material_unit
                material.supplier = get_object_or_404(Supplier, supplier_id=supplier_id)
                material.save()

                MaterialFinish.objects.filter(material=material).delete()

                finish_data = json.loads(request.POST.get("finish_data"))
                if finish_data:
                    for finish in finish_data:
                        finish_name = finish["finish_name"]
                        finish_cost = finish["finish_cost"]
                        if finish_name != 'delete':
                            new_finish = MaterialFinish.objects.create(finish_name=finish_name, finish_cost=finish_cost, material=material)

                response['url'] = reverse('material_list')  # URL to direct is str
                return JsonResponse(response)
    return render(request, 'cpq_app/material_detail.html', {'material_object': material_object, 'finishes': finishes, 'suppliers': suppliers, 'supplier_count': supplier_count})

def create_material(request):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    suppliers = Supplier.objects.all()
    supplier_count = len(suppliers) or 0
    if request.method == "POST":
        response = {}
        response['status'] = True
        print(request.POST)

        if request.POST.get("supplier_data"):
            for supplier in json.loads(request.POST.get("supplier_data")):
                print(supplier)
                supplier_operations(supplier)
            response['url'] = reverse('create_material')  # URL to direct is str
            print(response)
            return JsonResponse(response)
        else:
            material_name = request.POST.get("material_name")
            material_cost = request.POST.get("material_cost") or 0
            material_type = request.POST.get("material_type")
            material_unit = request.POST.get("material_unit")
            supplier_id = request.POST.get("supplier")
            supplier = Supplier.objects.get(supplier_id=supplier_id)

            new_material = Material.objects.create(material_name=material_name, 
            material_cost=material_cost, 
            material_type=material_type, 
            material_unit=material_unit, 
            supplier=supplier)

            finish_data = json.loads(request.POST.get("finish_data"))
            if finish_data:
                for finish in finish_data:
                    finish_name = finish["finish_name"]
                    finish_cost = finish["finish_cost"]
                    if finish_name != 'delete':
                        new_finish = MaterialFinish.objects.create(finish_name=finish_name, finish_cost=finish_cost, material=new_material)

            response['url'] = reverse('material_list')  # URL to direct is str
            print(response)
            return JsonResponse(response)
    return render(request, 'cpq_app/create_material.html', {'suppliers': suppliers, 'supplier_count': supplier_count})

def supplier_operations(supplier):

    supplier_id = supplier["supplier_id"]
    supplier_name = supplier["supplier_name"]
    if supplier_name == "delete":
        print('delete!')
        if supplier_id:
            Supplier.objects.filter(supplier_id=supplier_id).delete()
    elif supplier_id:
        print('edit!')
        supplier_object = Supplier.objects.get(supplier_id=supplier_id)
        supplier_object.supplier_name = supplier_name
        supplier_object.save()
    else:
        new_supplier = Supplier.objects.create(supplier_name = supplier_name)
    return True

def get_material_data_by_suppliers(suppliers, supplier_id = None):
    if supplier_id:
        supplier = Supplier.objects.get(supplier_id=supplier_id)
        material_data_by_suppliers = {
            "supplier_id": supplier.supplier_id,
            "supplier_name": supplier.supplier_name,
            "materials": []
        }
        supplier_materials = supplier.material_set.all()
        for material in supplier_materials:
            temp_material = {
                "material_id": material.material_id,
                "material_name": material.material_name,
                "material_type": material.material_type,
                "material_unit": material.material_unit,
                "material_cost": material.material_cost,
                "material_finishes": [],
            }
            finishes = material.materialfinish_set.all()
            for finish in finishes:
                temp_material["material_finishes"].append({
                    "finish_id": finish.finish_id,
                    "finish_name": finish.finish_name,
                    "finish_cost": finish.finish_cost,
                })
            material_data_by_suppliers["materials"].append(temp_material)
    else:
        material_data_by_suppliers = []
        for supplier in suppliers:
            temp_supplier = {
                "supplier_id": supplier.supplier_id,
                "supplier_name": supplier.supplier_name,
                "accessories": [],
                "glass": [],
                "aluminum": [],
            }

            supplier_materials = supplier.material_set.all()
            for material in supplier_materials:
                temp_material = {
                    "material_id": material.material_id,
                    "material_name": material.material_name,
                    "material_type": material.material_type,
                    "material_unit": material.material_unit,
                    "material_cost": material.material_cost,
                    "material_finishes": [],
                }

                finishes = material.materialfinish_set.all()
                for finish in finishes:
                    temp_material["material_finishes"].append({
                        "finish_id": finish.finish_id,
                        "finish_name": finish.finish_name,
                        "finish_cost": finish.finish_cost,
                    })

                # categorize based on material_type
                if material.material_type == "accessory":
                    temp_supplier["accessories"].append(temp_material)
                elif material.material_type == "glass":
                    temp_supplier["glass"].append(temp_material)
                elif material.material_type == "aluminum":
                    temp_supplier["aluminum"].append(temp_material)
            material_data_by_suppliers.append(temp_supplier)

    print(material_data_by_suppliers)
    return material_data_by_suppliers

def get_products(request):
    supplier_id = request.GET.get("supplier_id")  # Get supplier ID from AJAX request
    if supplier_id:
        supplier = get_object_or_404(Supplier, supplier_id=supplier_id)
        products = supplier.product_set.values("product_id", "product_name")  # Get product IDs and names
        print(list(products))
        return JsonResponse({"products": list(products)})  # Return JSON response
    return JsonResponse({"error": "Invalid request"}, status=400)

def delete_material(material_id):
    material_object = Material.objects.get(material_id=material_id)
    material_object.delete()

def get_customer(request):
    customer_id = request.GET.get('customer_id') 
    if customer_id:
        customer = get_object_or_404(Customer, customer_id=customer_id)  
        data = {
            "address": customer.customer_address,
            "mobile": customer.customer_mobile,
            "email": customer.customer_email
        }
        return JsonResponse(data) 
    return JsonResponse({"error": "Invalid request"}, status=400) 


def get_bill_of_materials(request):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    product_id = request.GET.get("product_id")
    height = float(request.GET.get("item_height"))
    width = float(request.GET.get("item_width"))
    item_quantity = float(request.GET.get("item_quantity"))
    excluded_materials = request.GET.getlist("excluded_materials[]")
    glass_finish = request.GET.get("glass_finish", "")
    aluminum_finish = request.GET.get("aluminum_finish", "")
    product_margin = float(request.GET.get("quotation_margin"))
    labor_margin = float(request.GET.get("quotation_labor"))
    print(request.GET)
    product = get_object_or_404(Product, product_id=product_id)
    materials = ProductMaterial.objects.filter(product=product)
    
    bom = []
    print(excluded_materials)
    for m in materials:
        mat_id = str(m.material.material_id)
        if mat_id in excluded_materials:
            continue

        mat = m.material
        finish = ""
        unit_cost = 0.0

        if mat.material_type == "accessory":
            single_unit = 1
            unit_cost = float(mat.material_cost)
        elif mat.material_type == "glass":
            finish_name = glass_finish
            finish_obj = MaterialFinish.objects.get(finish_name=finish_name, material=mat)
            unit_cost = float(finish_obj.finish_cost)
            finish = finish_name
            single_unit = (height*float(m.scale_ratio_second))*(width*float(m.scale_ratio))
        else:
            finish_name = aluminum_finish
            finish_obj = MaterialFinish.objects.get(finish_name=finish_name, material=mat)
            unit_cost = float(finish_obj.finish_cost)
            finish = finish_name
            dimension = height if m.scale_by_height else width
            single_unit = dimension * float(m.scale_ratio)

        single_item_qty = single_unit * float(m.material_quantity)
        single_unit_cost = single_unit * unit_cost
        single_item_cost = single_item_qty * unit_cost
        total_unit_qty = single_item_qty * item_quantity
        total_cost = total_unit_qty * unit_cost

        bom.append({
            "material_id": mat.material_id,
            "material_name": mat.material_name,
            "quantity": f"{m.material_quantity:.2f}",
            "material_unit": mat.material_unit.lower(),
            "material_single_unit": single_unit,
            "material_single_item_quantity": f"{single_item_qty:.2f}",
            "material_single_unit_cost": f"{single_unit_cost:.2f}",
            "material_single_item_quantity_cost": f"{single_item_cost:.2f}",
            "material_unit_total_quantity": f"{total_unit_qty:.2f}",
            "material_total_cost": f"{total_cost:.2f}",
            "material_finish": finish,
            "material_type": mat.material_type,
        })

    # Aggregate materials
    aggregated_bom = defaultdict(lambda: {
        "material_name": "",
        "quantity": 0.0,
        "material_unit": "",
        "material_single_unit": [],
        "material_single_item_quantity": 0.0,
        "material_single_unit_cost": [],
        "material_single_item_quantity_cost": 0.0,
        "material_unit_total_quantity": 0.0,
        "material_total_cost": 0.0,
        "material_finish": "",
        "material_type": "",
    })

    for item in bom:
        mat_id = item["material_id"]
        agg = aggregated_bom[mat_id]
        agg["material_name"] = item["material_name"]
        agg["quantity"] += float(item["quantity"])
        agg["material_unit"] = item["material_unit"]
        agg["material_single_unit"].append(item["material_single_unit"])
        agg["material_single_item_quantity"] += float(item["material_single_item_quantity"])
        agg["material_single_unit_cost"].append(float(item["material_single_unit_cost"]))
        agg["material_single_item_quantity_cost"] += float(item["material_single_item_quantity_cost"])
        agg["material_unit_total_quantity"] += float(item["material_unit_total_quantity"])
        agg["material_total_cost"] += float(item["material_total_cost"])
        agg["material_finish"] = item["material_finish"]
        agg["material_type"] = item["material_type"]

    # Prepare sorted result
    result_bom = sorted([
        {
            "material_id": mat_id,
            "material_name": data["material_name"],
            "quantity": f"{data['quantity']:.2f}",
            "material_unit": data["material_unit"],
            "material_single_unit": [f"{x:.2f}" for x in set(data["material_single_unit"])],
            "material_single_item_quantity": f"{data['material_single_item_quantity']:.2f}",
            "material_single_unit_cost": [f"{x:.2f}" for x in set(data["material_single_unit_cost"])],
            "material_single_item_quantity_cost": f"{data['material_single_item_quantity_cost']:.2f}",
            "material_unit_total_quantity": f"{data['material_unit_total_quantity']:.2f}",
            "material_total_cost": f"{data['material_total_cost']:.2f}",
            "material_finish": data["material_finish"],
            "material_type": data["material_type"],
        }
        for mat_id, data in aggregated_bom.items()
    ], key=lambda x: x["material_name"])

    additional_materials_raw = request.GET.get('additional_materials')
    additional_materials = json.loads(additional_materials_raw) if additional_materials_raw else []

    additional_bom = []
    for mat in additional_materials:
        material_id = mat['material_id']

        if not material_id or material_id == '[]': 
                continue
        quantity = float(mat['quantity'])
        finish_name = mat['finish']

        material = Material.objects.get(material_id = material_id)
        if finish_name:
            finish_obj = MaterialFinish.objects.get(finish_name=finish_name, material=material)
            cost = float(finish_obj.finish_cost) * quantity
        else:
            cost = float(material.material_cost) * quantity
        total_quantity = quantity*item_quantity
        per_item_cost_additional = item_quantity*cost
        additional_bom.append({
            'material_id': material_id,
            'material_name': material.material_name,
            'quantity': f"{quantity:.2f}",
            'finish_name': finish_name,
            'cost': f"{cost:.2f}",
            'total_quantity': f"{total_quantity:.2f}",
            'total_cost': f"{per_item_cost_additional:.2f}",
            'material_type': material.material_type,
            'material_unit': material.material_unit.lower(),
        })
    # Compute totals
    total_cost = sum(float(material["material_total_cost"]) for material in result_bom) + sum(float(material["total_cost"]) for material in additional_bom)
    per_item_cost = sum(float(material["material_single_item_quantity_cost"]) for material in result_bom) + sum(float(material["cost"]) for material in additional_bom)
    price_per_item = per_item_cost * (1 + (labor_margin + product_margin) / 100)
    total_price = price_per_item * item_quantity

    return JsonResponse({
        "product": product.product_name,
        "product_id": product.product_id,
        "bom": result_bom,
        "total_cost_of_materials": f"{total_cost:.2f}",
        "cost_of_materials_per_item": f"{per_item_cost:.2f}",
        "product_margin": product_margin,
        "labor_margin": labor_margin,
        "price_per_item": f"{price_per_item:.2f}",
        "total_price": f"{total_price:.2f}",
        "item_quantity": item_quantity,
        "item_id": request.GET.get('item_id'),
        "additional_bom": additional_bom,
    })



def get_total_bom(request):
    account_level = request.session.get("account_level")

    if account_level != "superuser":
        return redirect('access_error')

    print(request.POST)
    grand_bom = defaultdict(lambda: {
        "material_name": "",
        "quantity": 0.0,
        "material_unit": "",
        "material_single_unit": [],
        "material_single_item_quantity": 0.0,
        "material_single_unit_cost": [],
        "material_single_item_quantity_cost": 0.0,
        "material_unit_total_quantity": 0.0,
        "material_total_cost": 0.0,
        "material_finish": "",
        "material_type": "",
    })

    total_cost_of_materials = 0
    cost_of_materials_per_item = 0
    total_price = 0
    total_quantity = 0
    last_product_id = None
    last_margin = 0
    last_labor = 0
    data = json.loads(request.POST.get("submit_data", "[]"))
    total_item_quantity = 0

    additional_bom = []
    for item in data:
        product_id = item["product_id"]
        height = float(item["item_height"])
        width = float(item["item_width"])
        item_quantity = float(item["item_quantity"])
        excluded_materials = item.get("excluded_materials", [])
        glass_finish = item.get("glass_finish", "")
        aluminum_finish = item.get("aluminum_finish", "")
        product_margin = float(item.get("submit_margin", 0))
        labor_margin = float(item.get("submit_labor", 0))

        product = get_object_or_404(Product, product_id=product_id)
        materials = ProductMaterial.objects.filter(product=product)

        for material in materials:
            if str(material.material.material_id) in excluded_materials:
                continue

            material_quantity = float(material.material_quantity)
            scale_by_height = material.scale_by_height
            scale_by_width = material.scale_by_width
            scale_ratio = float(material.scale_ratio)
            material_obj = material.material
            material_type = material_obj.material_type

            # Determine cost and quantity
            if material_type == "accessory":
                material_finish = ""
                cost = float(material_obj.material_cost)
                material_single_unit = 1
            elif material_type =="glass":
                material_finish = glass_finish
                finish_object = MaterialFinish.objects.get(finish_name=material_finish, material=material_obj)
                cost = float(finish_object.finish_cost)
                scale_ratio_second = float(material.scale_ratio_second)
                material_single_unit = (height*float(scale_ratio_second))*(width*float(scale_ratio))
            else:
                material_finish = aluminum_finish
                finish_object = MaterialFinish.objects.get(finish_name=material_finish, material=material_obj)
                cost = float(finish_object.finish_cost)
                material_single_unit = height * scale_ratio if scale_by_height else width * scale_ratio

            material_single_item_quantity = material_quantity * material_single_unit
            material_single_unit_cost = material_single_unit * cost
            material_single_item_quantity_cost = material_single_item_quantity * cost
            material_unit_total_quantity = material_single_item_quantity * item_quantity
            material_total_cost = material_unit_total_quantity * cost

            # Composite key: material_id + material_finish
            mat_key = (material_obj.material_id, material_finish)
            bom_entry = grand_bom[mat_key]
            bom_entry["material_name"] = material_obj.material_name
            bom_entry["material_unit"] = material_obj.material_unit.lower()
            bom_entry["material_finish"] = material_finish
            bom_entry["material_type"] = material_type
            bom_entry["quantity"] += material_quantity
            bom_entry["material_single_unit"].append(material_single_unit)
            bom_entry["material_single_item_quantity"] += material_single_item_quantity
            bom_entry["material_single_unit_cost"].append(material_single_unit_cost)
            bom_entry["material_single_item_quantity_cost"] += material_single_item_quantity_cost
            bom_entry["material_unit_total_quantity"] += material_unit_total_quantity
            bom_entry["material_total_cost"] += material_total_cost

        # Recalculate total costs only for non-excluded materials
        for material in materials:
            if str(material.material.material_id) in excluded_materials:
                continue
            material_quantity = float(material.material_quantity)
            scale_ratio = float(material.scale_ratio)
            scale_by_height = material.scale_by_height
            scale_by_width = material.scale_by_width
            material_obj = material.material
            material_type = material_obj.material_type

            if material_type == "accessory":
                cost = float(material_obj.material_cost)
                single_unit = 1
            else:
                finish_name = glass_finish if material_type == "glass" else aluminum_finish
                finish_object = MaterialFinish.objects.get(finish_name=finish_name, material=material_obj)
                cost = float(finish_object.finish_cost)
                single_unit = height * scale_ratio if scale_by_height else width * scale_ratio

            cost_of_materials_per_item += material_quantity * cost * single_unit
            total_cost_of_materials += material_quantity * single_unit * item_quantity * cost

        result_bom = sorted([
            {
                "material_id": mat_key[0],
                "material_finish": mat_key[1],
                "material_name": values["material_name"],
                "quantity": f"{values['quantity']:.2f}",
                "material_unit": values["material_unit"],
                "material_single_unit": [f"{x:.2f}" for x in set(values["material_single_unit"])],
                "material_single_item_quantity": f"{values['material_single_item_quantity']:.2f}",
                "material_single_unit_cost": [f"{x:.2f}" for x in set(values["material_single_unit_cost"])],
                "material_single_item_quantity_cost": f"{values['material_single_item_quantity_cost']:.2f}",
                "material_unit_total_quantity": f"{values['material_unit_total_quantity']:.2f}",
                "material_total_cost": f"{values['material_total_cost']:.2f}",
                "material_type": values["material_type"],
            }
            for mat_key, values in grand_bom.items()
        ], key=lambda x: x["material_name"])
        
        additional_materials_raw = item['additional_materials']
        additional_materials = json.loads(additional_materials_raw) if additional_materials_raw else []

        print(additional_materials)
        for mat in additional_materials:
            material_id = mat['material_id']
            print(material_id)
            if not material_id or material_id == '[]': 
                continue
            quantity = float(mat['quantity'])
            finish_name = mat['finish']

            material = Material.objects.get(material_id=material_id)
            
            if finish_name:
                finish_obj = MaterialFinish.objects.get(finish_name=finish_name, material=material)
                cost = float(finish_obj.finish_cost)* quantity
            else:
                cost = float(material.material_cost)* quantity

            total_quantity = quantity * item_quantity
            per_item_cost_additional = item_quantity * cost

            additional_bom.append({
                'material_id': material_id,
                'material_name': material.material_name,
                'quantity': f"{quantity:.2f}",
                'finish_name': finish_name,
                'cost': f"{cost:.2f}",
                'total_quantity': f"{total_quantity:.2f}",
                'total_cost': f"{per_item_cost_additional:.2f}",
                'material_type': material.material_type,
                'material_unit': material.material_unit.lower(),
            })

            total_cost_of_materials += per_item_cost_additional

        total_item_quantity+= item_quantity
    
    total_cost = sum(float(item["material_total_cost"]) for item in result_bom) + sum(float(material["total_cost"]) for material in additional_bom)
    total_price = total_cost * (1 + (labor_margin + product_margin) / 100)
    print(additional_bom)
    response_data = {
        "bom": result_bom,
        "total_cost_of_materials": f"{total_cost:.2f}",
        "product_margin": product_margin,
        "labor_margin": labor_margin,
        "total_price": f"{total_price:.2f}",
        "item_quantity": total_item_quantity,
        "additional_bom": additional_bom,
    }

    return JsonResponse(response_data)
def download_quotation_excel(request, quotation_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    quotation = Quotation.objects.get(quotation_id=quotation_id)

    center_alignment = Alignment(horizontal="center")
    # Title and Company Header
    ws.merge_cells('A1:I1')
    ws['A1'] = "CRIVER GLASS AND ALUMINUM CORP."
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:I2')
    ws['A2'] = "Unit 1-B GV Square Casa Mila Subd. Commonwealth Ave. Ext., Quezon City"
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:I3')
    ws['A3'] = "Contact nos.: 632.897.2303   Fascimile no.: 632.990.5064"
    ws['A3'].alignment = center_alignment

    ws.merge_cells('A4:I4')
    ws['A4'] = "E-mail address: criver.glassandalum@gmail.com"
    ws['A4'].alignment = center_alignment

    # Quotation Details
    ws.merge_cells('A6:I6')
    ws['A6'] = "QUOTATION CONTRACT"
    ws['A6'].font = Font(bold=True, size=14)
    ws['A6'].alignment = center_alignment

    ws.merge_cells('A8:B8')
    ws['A8'] = "Name of Client:"
    ws.merge_cells('C8:F8')
    ws['C8'] = quotation.customer.customer_name

    ws.merge_cells('G8:H8')
    ws['G8'] = "Date:"
    ws['I8'] = datetime.today().date()

    ws.merge_cells('A9:B9')
    ws['A9'] = "Project:"
    ws.merge_cells('C9:F9')
    ws['C9'] = quotation.project

    ws.merge_cells('G9:H9')
    ws['G9'] = "Reference No.:"
    ws['I9'] = "2025.001"

    ws.merge_cells('A10:B10')
    ws['A10'] = "Address:"
    ws.merge_cells('C10:F10')
    ws['C10'] = quotation.customer.customer_address

    ws.merge_cells('G10:H10')
    ws['G10'] = "Contact No.:"
    ws['I10'] = quotation.customer.customer_mobile

    ws.append([])
    # Table Header
    headers = ["QTY", "UNIT", "DESCRIPTION", "", "", "", "", "UNIT PRICE", "AMOUNT"]
    ws.append(headers)
    ws.merge_cells('C12:G12')
    for cell in ["A12", "B12", "C12", "H12", "I12"]:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = center_alignment

    # Sample data rows (you can dynamically loop your queryset here)
    ws.append([])
    ws.append([])

    sum_total_price = 0
    for item in QuotationItem.objects.filter(quotation=quotation):
        unit_price = item.get_unit_price()
        quantity = item.item_quantity
        total_price = unit_price*quantity
        ws.append(["", "", "", "", "", "", "", "", ""])  # Placeholder for spacing
        label_row = ws.max_row
        ws.cell(row=label_row, column=3, value=f"{item.item_label}").font = Font(bold=True)
        detail_row = [
            quantity,
            "sets",
            f"{item.item_width} (w)",
            "X",
            f"{item.item_height} (h)",
            f"{item.product.product_name}",
            "",
            unit_price,
            total_price
        ]
        ws.append(detail_row)

        detail_row_num = ws.max_row
        for col in [1, 2, 3, 4, 5]: 
            ws.cell(row=detail_row_num, column=col).alignment = Alignment(horizontal="center")

        sum_total_price += total_price
    table_end_row = ws.max_row
    last_data_row = ws.max_row + 1
    # TOTAL
    ws[f'G{last_data_row}'] = "TOTAL:"
    ws[f'I{last_data_row}'] = sum_total_price
    ws[f'G{last_data_row}'].font = Font(bold=True)

    # V.A.T.
    last_data_row += 1
    ws[f'G{last_data_row}'] = "V.A.T.:"
    ws[f'I{last_data_row}'] = ""
    ws[f'G{last_data_row}'].font = Font(bold=True)

    # GRAND TOTAL
    last_data_row += 1
    ws.merge_cells(f'F{last_data_row}:G{last_data_row}')
    ws[f'F{last_data_row}'] = "GRAND TOTAL COST"
    ws[f'I{last_data_row}'] = sum_total_price
    ws[f'F{last_data_row}'].font = Font(bold=True)

    start_row = 12 
    end_row = last_data_row 

    
    thin = Side(style="thin")

    for row in range(start_row, end_row + 1):  # Iterating from row 12 to end_row
        for col in range(1, 10):  # Columns A to I
            cell = ws.cell(row=row, column=col)

            # Determine cell position
            is_first_row = (row == start_row)
            is_last_row = (row == end_row)
            is_table_end_row = (row == table_end_row)
            is_col_1 = (col == 1)
            is_col_9 = (col == 9)
            col_letter = chr(64 + col)  # Convert 1 -> A, 2 -> B, ..., 9 -> I

            # Determine if the column needs side borders (A, B, H, I)
            needs_side_borders = col_letter in ['A', 'B', 'H', 'I']
            is_row_12 = (row == 12)

            # Initialize sides to None
            top = bottom = left = right = None

            # Outer border logic
            if is_first_row:
                top = thin
            if is_last_row or is_table_end_row:
                bottom = thin
            if is_col_1:
                left = thin
            if is_col_9:
                right = thin

            # Special full borders for A12 to I12 (row 12)
            if is_row_12:
                top = bottom = left = right = thin

            # Apply left/right borders to A, B, H, I from row 12 onwards
            if needs_side_borders and row >= 12:
                left = thin
                right = thin
                if row == 12:
                    top = thin
                if row == table_end_row:
                    bottom = thin

            # **Check for the bottom-right most cell (I/end_row) and apply a full border**
            if row == end_row and col == 9:  # Cell I in the last row
                top = bottom = left = right = thin

            # Assign the combined border
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)




    # Set column widths for better formatting
    ws.column_dimensions['A'].width = 9
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 11
    ws.column_dimensions['I'].width = 12


    # Return Excel response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=quotation.xlsx'
    wb.save(response)
    return response

# misc for the faq and about
def faq(request):
    if not is_logged_in(request):
        return redirect('login')

    return render(request, 'cpq_app/faq.html')

def about(request):
    if not is_logged_in(request):
        return redirect('login')

    return render(request, 'cpq_app/about.html')

def feedback(request):
    if not is_logged_in(request):
        return redirect('login')

    return render(request, 'cpq_app/feedback.html')


#DELETE THIS LATER!
'''
def copy_aluminum_materials(new_supplier_id, finish_cost_increase):
    # Get the new supplier instance
    new_supplier = Supplier.objects.get(supplier_id=new_supplier_id)
    roosevelt = Supplier.objects.get(supplier_id=13)

    # Get all aluminum materials
    aluminum_materials = Material.objects.filter(material_type='aluminum', supplier=roosevelt)

    for material in aluminum_materials:
        # Create a copy of the aluminum material with a new supplier
        new_material = Material(
            material_name=material.material_name,
            material_type=material.material_type,
            material_unit=material.material_unit,
            supplier=new_supplier
        )
        new_material.save()

        # Get the material finishes for the original material
        material_finishes = MaterialFinish.objects.filter(material=material)

        for finish in material_finishes:
            # Create a copy of the material finish with an updated finish cost
            new_finish = MaterialFinish(
                finish_name=finish.finish_name,
                finish_cost=float(finish.finish_cost) * finish_cost_increase,
                material=new_material
            )
            new_finish.save()

    print("Aluminum materials and associated finishes copied successfully!")
'''
